#!/usr/bin/env python3

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_PATH = ROOT / "data" / "DRVs_All_populations.xlsx"
OUTPUT_PATH = ROOT / "data" / "drvs_eu_normalized.json"

REFERENCE_FIELDS = [
    ("AI", "AI"),
    ("AR", "AR"),
    ("PRI", "PRI"),
    ("RI", "RI"),
    ("UL", "UL"),
    ("Safe and adequate intake", "safe_and_adequate_intake"),
]


def parse_age_label(age_label):
    if age_label is None:
        return {"raw": None, "min": None, "max": None, "unit": None, "comparator": None, "pal": None}

    raw = str(age_label).strip()
    parsed = {"raw": raw, "min": None, "max": None, "unit": None, "comparator": None, "pal": None}

    pal_match = re.search(r"PAL\s*=\s*([0-9.]+)", raw)
    if pal_match:
        parsed["pal"] = float(pal_match.group(1))
        raw = raw[: pal_match.start()].strip()

    raw = raw.replace("≥", ">=").replace("–", "-")

    gte_match = re.match(r"^>=\s*([0-9.]+)\s*(month|months|year|years)$", raw, re.IGNORECASE)
    if gte_match:
        parsed["min"] = float(gte_match.group(1))
        parsed["unit"] = "months" if "month" in gte_match.group(2).lower() else "years"
        parsed["comparator"] = ">="
        return parsed

    range_match = re.match(
        r"^([0-9.]+)\s*-\s*([0-9.]+)\s*(month|months|year|years)$", raw, re.IGNORECASE
    )
    if range_match:
        parsed["min"] = float(range_match.group(1))
        parsed["max"] = float(range_match.group(2))
        parsed["unit"] = "months" if "month" in range_match.group(3).lower() else "years"
        return parsed

    single_match = re.match(r"^([0-9.]+)\s*(month|months|year|years)$", raw, re.IGNORECASE)
    if single_match:
        parsed["min"] = float(single_match.group(1))
        parsed["max"] = float(single_match.group(1))
        parsed["unit"] = "months" if "month" in single_match.group(2).lower() else "years"
        return parsed

    return parsed


def parse_reference_value(value):
    if value is None:
        return None

    raw = str(value).strip()
    if not raw:
        return None

    normalized = raw.replace("–", "-")
    if normalized in {"ND", "ND.", "ND ", "NA", "NA.", "NA. "}:
        return {"raw": raw, "status": normalized.rstrip(". "), "kind": "not_available"}

    range_match = re.match(r"^([0-9.]+)\s*-\s*([0-9.]+)\s+(.+)$", normalized)
    if range_match:
        return {
            "raw": raw,
            "status": "value",
            "kind": "range",
            "min": float(range_match.group(1)),
            "max": float(range_match.group(2)),
            "unit": range_match.group(3).strip(),
        }

    value_match = re.match(r"^([0-9.]+)\s+(.+)$", normalized)
    if value_match:
        return {
            "raw": raw,
            "status": "value",
            "kind": "scalar",
            "value": float(value_match.group(1)),
            "unit": value_match.group(2).strip(),
        }

    return {"raw": raw, "status": "value", "kind": "text"}


def transform():
    wb = load_workbook(INPUT_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        (
            category,
            nutrient,
            target_population,
            age,
            gender,
            ai,
            ar,
            pri,
            ri,
            ul,
            safe_and_adequate_intake,
        ) = row

        references = {}
        for raw_name, output_name in REFERENCE_FIELDS:
            references[output_name] = parse_reference_value(locals()[raw_name.lower().replace(" ", "_")])

        records.append(
            {
                "source_system": "EFSA",
                "source_file": INPUT_PATH.name,
                "category": category,
                "nutrient": nutrient,
                "target_population": target_population,
                "gender": gender,
                "age": parse_age_label(age),
                "references": references,
            }
        )

    payload = {
        "dataset": "EU Dietary Reference Values",
        "source_system": "EFSA",
        "source_file": INPUT_PATH.name,
        "record_count": len(records),
        "records": records,
    }

    OUTPUT_PATH.write_text(
    json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
    encoding="utf-8"
    )
    print(f"Wrote {len(records)} records to {OUTPUT_PATH}")


if __name__ == "__main__":
    transform()
