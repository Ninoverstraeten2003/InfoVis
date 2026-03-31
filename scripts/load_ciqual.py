#!/usr/bin/env python3
"""
Load Ciqual food composition data into canonical tables.

Populates: source_file, source_row, food, food_nutrient_value.

Idempotent: re-running clears previous CIQUAL source_rows (cascading to
food → food_nutrient_value) before re-inserting.

Prerequisites:
    - Schema created (db/schema.sql)
    - Nutrients + aliases seeded (scripts/seed_canonical.py)

Usage:
    python scripts/load_ciqual.py
"""

import json
import os
import re
from pathlib import Path

import psycopg2
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT_PATH = ROOT / "data" / "Table Ciqual 2025_ENG_2025_11_03.xlsx"

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://localhost:5432/nutriverse",
)

# Identity column indices (0-based) in the Ciqual "food composition" sheet
IDENTITY_COLS = {
    0: "alim_grp_code",
    1: "alim_ssgrp_code",
    2: "alim_ssssgrp_code",
    3: "alim_grp_nom_eng",
    4: "alim_ssgrp_nom_eng",
    5: "alim_ssssgrp_nom_eng",
    6: "alim_code",
    7: "alim_nom_eng",
    8: "alim_nom_sci",
}

# First nutrient column index (0-based)
NUTRIENT_START_COL = 9

# Skip the "Jones factor" column — it's metadata, not a nutrient value
SKIP_COLUMNS = {"Jones factor"}


def parse_ciqual_value(raw_cell):
    """Parse a Ciqual cell value into (numeric_value, quality_flag).

    Returns:
        (value, quality_flag) where:
            value: float or None
            quality_flag: 'measured', 'less_than', 'traces', 'not_analyzed', 'missing'
    """
    if raw_cell is None:
        return None, "missing"

    raw = str(raw_cell).strip()
    if not raw:
        return None, "missing"

    if raw == "-":
        return None, "not_analyzed"

    if raw.lower() in ("traces", "trace"):
        return 0.0, "traces"

    # Handle "< X,XX" or "< X.XX"
    lt_match = re.match(r"^<\s*([0-9]+[.,]?[0-9]*)$", raw)
    if lt_match:
        val_str = lt_match.group(1).replace(",", ".")
        return float(val_str), "less_than"

    # Handle normal numeric (comma as decimal separator)
    cleaned = raw.replace(",", ".").strip()
    try:
        return float(cleaned), "measured"
    except ValueError:
        # Unparseable — store raw, flag as missing
        return None, "missing"


def extract_unit_from_header(header):
    """Extract unit from the LAST set of parentheses in the header."""
    # findall gets all matches: ['vitamine E', 'mg 100g']
    matches = re.findall(r"\(([^)]+)\)", header)
    if matches:
        # Take the last one: 'mg 100g'
        inside = matches[-1]
        parts = inside.strip().split()
        if len(parts) >= 1:
            return parts[0]
    return None


def clean_header(raw_header):
    """Clean newlines from Ciqual headers for storage."""
    if raw_header is None:
        return None
    return str(raw_header).replace("\n", " ").strip()


def load():
    print(f"Reading {INPUT_PATH.name}...")
    wb = load_workbook(INPUT_PATH, read_only=True, data_only=True)
    ws = wb["food composition"]

    # Read headers
    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [clean_header(h) for h in raw_headers]

    # Build nutrient column map: col_index → cleaned header
    nutrient_columns = {}
    for i in range(NUTRIENT_START_COL, len(headers)):
        h = headers[i]
        if h and h not in SKIP_COLUMNS:
            nutrient_columns[i] = h

    print(f"Found {len(nutrient_columns)} nutrient columns")

    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor() as cur:
            # Source file
            cur.execute(
                """
                INSERT INTO source_file (source_system, file_name, dataset_name)
                VALUES (%s, %s, %s)
                ON CONFLICT (source_system, file_name)
                DO UPDATE SET dataset_name = EXCLUDED.dataset_name
                RETURNING id
                """,
                ("CIQUAL", INPUT_PATH.name, "Ciqual French Food Composition 2025"),
            )
            source_file_id = cur.fetchone()[0]

            # Clear previous load for idempotence
            # CASCADE: source_row → food (via source_row_id) → food_nutrient_value (via food_id)
            cur.execute(
                "DELETE FROM source_row WHERE source_file_id = %s",
                (source_file_id,),
            )
            cleared = cur.rowcount
            if cleared > 0:
                print(f"  Cleared {cleared} previous source_row(s) (+ cascaded food/food_nutrient_value)")

            # Also clear orphaned food rows that lost their source_row_id
            # (food has ON DELETE CASCADE from source_row, so the above handles it)

            # Build alias → nutrient_id lookup for CIQUAL source
            cur.execute(
                """
                SELECT na.alias, na.nutrient_id
                FROM nutrient_alias na
                WHERE na.source_system = 'CIQUAL'
                """
            )
            alias_map = dict(cur.fetchall())
            print(f"Loaded {len(alias_map)} CIQUAL alias mappings")

            # Map each column header to a nutrient_id
            col_nutrient_map = {}
            unmapped = []
            for col_idx, col_header in nutrient_columns.items():
                nutrient_id = alias_map.get(col_header)
                if nutrient_id is not None:
                    col_nutrient_map[col_idx] = (nutrient_id, col_header)
                else:
                    unmapped.append(col_header)

            if unmapped:
                print(f"WARNING: {len(unmapped)} Ciqual columns have no nutrient alias mapping:")
                for h in unmapped:
                    print(f"  - {h}")
                print("These columns will be skipped during loading.")

            print(f"Mapped {len(col_nutrient_map)} columns to canonical nutrients")

            food_inserted = 0
            food_updated = 0
            fnv_inserted = 0
            fnv_updated = 0
            fnv_skipped = 0

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row = list(row)
                if all(cell is None for cell in row):
                    continue

                # Extract identity columns
                alim_code = str(row[6]) if row[6] is not None else None
                alim_nom = str(row[7]) if row[7] is not None else None
                if not alim_code or not alim_nom:
                    continue

                # Store source row (full row as JSONB)
                raw_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        raw_dict[headers[i]] = str(val) if val is not None else None
                cur.execute(
                    """
                    INSERT INTO source_row (source_file_id, sheet_name, row_number, raw_payload)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (source_file_id, sheet_name, row_number)
                    DO UPDATE SET raw_payload = EXCLUDED.raw_payload
                    RETURNING id
                    """,
                    (source_file_id, "food composition", row_num, json.dumps(raw_dict)),
                )
                source_row_id = cur.fetchone()[0]

                # Insert food
                cur.execute(
                    """
                    INSERT INTO food
                        (source_row_id, source_food_code, name, scientific_name,
                         group_code, group_name,
                         subgroup_code, subgroup_name,
                         subsubgroup_code, subsubgroup_name)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (source_food_code) DO UPDATE
                        SET name = EXCLUDED.name,
                            source_row_id = EXCLUDED.source_row_id
                    RETURNING id, (xmax = 0) AS inserted
                    """,
                    (
                        source_row_id,
                        alim_code,
                        alim_nom.replace("\n", " ").strip(),
                        str(row[8]).strip() if row[8] else None,
                        str(row[0]) if row[0] else None,
                        str(row[3]).strip() if row[3] else None,
                        str(row[1]) if row[1] else None,
                        str(row[4]).strip() if row[4] else None,
                        str(row[2]) if row[2] else None,
                        str(row[5]).strip() if row[5] else None,
                    ),
                )
                food_id, food_was_insert = cur.fetchone()
                if food_was_insert:
                    food_inserted += 1
                else:
                    food_updated += 1

                # Insert nutrient values
                for col_idx, (nutrient_id, col_header) in col_nutrient_map.items():
                    raw_cell = row[col_idx] if col_idx < len(row) else None
                    value, quality_flag = parse_ciqual_value(raw_cell)
                    unit = extract_unit_from_header(col_header)

                    cur.execute(
                        """
                        INSERT INTO food_nutrient_value
                            (food_id, nutrient_id, value, unit, basis,
                             raw_column_name, raw_cell_value, quality_flag)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (food_id, nutrient_id, raw_column_name)
                        DO UPDATE SET
                            value          = EXCLUDED.value,
                            unit           = EXCLUDED.unit,
                            raw_cell_value = EXCLUDED.raw_cell_value,
                            quality_flag   = EXCLUDED.quality_flag
                        RETURNING (xmax = 0) AS inserted
                        """,
                        (
                            food_id, nutrient_id, value, unit, "per_100g",
                            col_header,
                            str(raw_cell) if raw_cell is not None else None,
                            quality_flag,
                        ),
                    )
                    if cur.fetchone()[0]:
                        fnv_inserted += 1
                    else:
                        fnv_updated += 1

                if (food_inserted + food_updated) % 500 == 0:
                    print(f"  ...processed {food_inserted + food_updated} foods")

            conn.commit()
            total_foods = food_inserted + food_updated
            total_fnv = fnv_inserted + fnv_updated
            print(f"\nfood: {food_inserted} inserted, {food_updated} updated ({total_foods} total)")
            print(f"food_nutrient_value: {fnv_inserted} inserted, {fnv_updated} updated ({total_fnv} total)")

    finally:
        conn.close()


if __name__ == "__main__":
    load()
