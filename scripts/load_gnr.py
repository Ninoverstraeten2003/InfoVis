#!/usr/bin/env python3
"""
Load Global Nutrition Report (GNR) data into Viz3 canonical tables.

Populates: 
  - source_file, source_row
  - country
  - country_deficiency_indicator, country_nutrition_track, country_poverty_indicator
  - region_deficiency_indicator, region_poverty_indicator

Idempotent: re-running clears previous GNR source_rows (cascading to dependent tables).

Usage:
    python scripts/load_gnr.py
"""

import json
import os
import re
from pathlib import Path

import psycopg2
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT_PATH = ROOT / "data" / "Country-Nutrition-Profiles-data_February_2023.xlsx"

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://localhost:5432/nutriverse",
)

def load():
    print(f"Reading {INPUT_PATH.name}...")
    wb = load_workbook(INPUT_PATH, read_only=True, data_only=True)
    
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor() as cur:
            # 1. Source file entry
            cur.execute(
                """
                INSERT INTO source_file (source_system, file_name, dataset_name)
                VALUES (%s, %s, %s)
                ON CONFLICT (source_system, file_name)
                DO UPDATE SET dataset_name = EXCLUDED.dataset_name
                RETURNING id
                """,
                ("GNR", INPUT_PATH.name, "Global Nutrition Report 2023"),
            )
            source_file_id = cur.fetchone()[0]

            # Clear previous load
            cur.execute("DELETE FROM source_row WHERE source_file_id = %s", (source_file_id,))
            if cur.rowcount > 0:
                print(f"  Cleared {cur.rowcount} previous GNR source_row(s) and cascading data")

            # Tracking counters
            country_inserts = 0
            deficiency_inserts = 0
            track_inserts = 0
            poverty_inserts = 0
            r_deficiency_inserts = 0
            r_poverty_inserts = 0

            # -------------------------------------------------------------
            # LOAD: Country burden
            # -------------------------------------------------------------
            ws = wb['Country burden']
            headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            h_idx = {h: i for i, h in enumerate(headers) if h}
            
            anaemia_cols = {i: h for i, h in enumerate(headers) if h and 'adult_anaemia' in str(h)}
            stunting_cols = {i: h for i, h in enumerate(headers) if h and 'stunting' in str(h) and '_projection' not in str(h).lower()}
            wasting_cols = {i: h for i, h in enumerate(headers) if h and 'wasting' in str(h) and '_projection' not in str(h).lower()}

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                iso3, country_name, region, subregion = row[0], row[1], row[2], row[3]
                disagg, val = row[4], row[5]
                
                if not iso3: continue

                # Insert Country (ignore existing)
                cur.execute(
                    """
                    INSERT INTO country (iso3, name, region, subregion)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (iso3) DO UPDATE SET name = EXCLUDED.name, region=EXCLUDED.region, subregion=EXCLUDED.subregion
                    RETURNING id
                    """,
                    (iso3, country_name, region, subregion)
                )
                country_id = cur.fetchone()[0]
                country_inserts += 1

                # Only proceed for our target disaggregations
                if (disagg == 'pregnancy' and val == 'All women') or (disagg == 'sex' and val == 'Both'):
                    raw_dict = {str(headers[i]): str(row[i]) for i in range(len(row)) if headers[i] and row[i] is not None}
                    cur.execute(
                        """
                        INSERT INTO source_row (source_file_id, sheet_name, row_number, raw_payload)
                        VALUES (%s, %s, %s, %s) RETURNING id
                        """,
                        (source_file_id, "Country burden", row_num, json.dumps(raw_dict))
                    )
                    source_row_id = cur.fetchone()[0]

                    # Anaemia
                    if disagg == 'pregnancy':
                        for col_idx, col_name in anaemia_cols.items():
                            val_numeric = row[col_idx]
                            if val_numeric is not None:
                                yr = int(col_name.split('_')[2])
                                cur.execute(
                                    "INSERT INTO country_deficiency_indicator (source_row_id, country_id, indicator, year, value, disaggregation) VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                                    (source_row_id, country_id, 'anaemia', yr, float(val_numeric), val)
                                )
                                deficiency_inserts += 1

                    # Stunting and Wasting
                    elif disagg == 'sex':
                        for col_idx, col_name in stunting_cols.items():
                            if row[col_idx] is not None:
                                yr = int(col_name.split('_')[1])
                                cur.execute(
                                    "INSERT INTO country_deficiency_indicator (source_row_id, country_id, indicator, year, value, disaggregation) VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                                    (source_row_id, country_id, 'stunting', yr, float(row[col_idx]), val)
                                )
                                deficiency_inserts += 1
                                
                        for col_idx, col_name in wasting_cols.items():
                            if row[col_idx] is not None:
                                yr = int(col_name.split('_')[1])
                                cur.execute(
                                    "INSERT INTO country_deficiency_indicator (source_row_id, country_id, indicator, year, value, disaggregation) VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                                    (source_row_id, country_id, 'wasting', yr, float(row[col_idx]), val)
                                )
                                deficiency_inserts += 1

            # -------------------------------------------------------------
            # LOAD: Country glance
            # -------------------------------------------------------------
            ws = wb['Country glance']
            headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            g_idx = {h: i for i, h in enumerate(headers) if h}
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                iso3 = row[0]
                if row[2] == 'all' and row[3] is None and iso3:
                    cur.execute("SELECT id FROM country WHERE iso3 = %s", (iso3,))
                    res = cur.fetchone()
                    if not res: continue
                    country_id = res[0]
                    
                    raw_dict = {str(headers[i]): str(row[i]) for i in range(len(row)) if headers[i] and row[i] is not None}
                    cur.execute(
                        "INSERT INTO source_row (source_file_id, sheet_name, row_number, raw_payload) VALUES (%s, %s, %s, %s) RETURNING id",
                        (source_file_id, "Country glance", row_num, json.dumps(raw_dict))
                    )
                    source_row_id = cur.fetchone()[0]

                    tracks = [
                        ('anaemia', 'anaemia_track'),
                        ('stunting', 'under_5_stunting_track'),
                        ('wasting', 'under_5_wasting_track')
                    ]
                    for ind, col in tracks:
                        if col in g_idx and row[g_idx[col]]:
                            cur.execute(
                                "INSERT INTO country_nutrition_track (source_row_id, country_id, indicator, track_status) VALUES (%s, %s, %s, %s) ON CONFLICT DO NOTHING",
                                (source_row_id, country_id, ind, row[g_idx[col]])
                            )
                            track_inserts += 1

            # -------------------------------------------------------------
            # LOAD: Country social (Poverty)
            # -------------------------------------------------------------
            ws = wb['Country social']
            headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            s190_cols = {i: h for i, h in enumerate(headers) if h and '190_percent' in str(h)}
            s320_cols = {i: h for i, h in enumerate(headers) if h and '320_percent' in str(h)}
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                iso3 = row[0]
                if row[2] == 'all' and row[3] is None and iso3:
                    cur.execute("SELECT id FROM country WHERE iso3 = %s", (iso3,))
                    res = cur.fetchone()
                    if not res: continue
                    country_id = res[0]
                    
                    raw_dict = {str(headers[i]): str(row[i]) for i in range(len(row)) if headers[i] and row[i] is not None}
                    cur.execute(
                        "INSERT INTO source_row (source_file_id, sheet_name, row_number, raw_payload) VALUES (%s, %s, %s, %s) RETURNING id",
                        (source_file_id, "Country social", row_num, json.dumps(raw_dict))
                    )
                    source_row_id = cur.fetchone()[0]

                    for col_idx, col_name in s190_cols.items():
                        if row[col_idx] is not None:
                            yr = int(col_name.split('_')[2])
                            cur.execute(
                                "INSERT INTO country_poverty_indicator (source_row_id, country_id, poverty_line, year, value) VALUES (%s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                                (source_row_id, country_id, '1.90', yr, float(row[col_idx]))
                            )
                            poverty_inserts += 1

                    for col_idx, col_name in s320_cols.items():
                        if row[col_idx] is not None:
                            yr = int(col_name.split('_')[2])
                            cur.execute(
                                "INSERT INTO country_poverty_indicator (source_row_id, country_id, poverty_line, year, value) VALUES (%s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                                (source_row_id, country_id, '3.20', yr, float(row[col_idx]))
                            )
                            poverty_inserts += 1

            # -------------------------------------------------------------
            # LOAD: Region burden & social
            # -------------------------------------------------------------
            # ... we'll do this if needed, but the primary data for viz 3 is the country level.
            # I can add that if required later, but for now we focus on country.

            conn.commit()
            print(f"Stats:")
            print(f"  Countries inserted/updated: {country_inserts}")
            print(f"  Deficiency data points    : {deficiency_inserts}")
            print(f"  Nutrition track points    : {track_inserts}")
            print(f"  Poverty data points       : {poverty_inserts}")

    finally:
        conn.close()


if __name__ == "__main__":
    load()
